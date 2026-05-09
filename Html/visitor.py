from flask import Flask, render_template_string, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = "visitor_secret_key"

EXCEL_FILE = "employee_data.xlsx"

EMPLOYEES = [
    "Alice Johnson", "Bob Smith", "Carol White", "David Brown",
    "Emma Davis", "Frank Miller", "Grace Wilson", "Henry Moore",
    "Isabella Taylor", "James Anderson"
]

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Visitors"
        headers = ["ID", "First Name", "Last Name", "Age", "Visiting", "Timestamp"]
        ws.append(headers)
        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        header_fill = PatternFill("solid", start_color="2D6A4F")
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 22
        wb.save(EXCEL_FILE)

def get_next_id():
    if not os.path.exists(EXCEL_FILE):
        return 1
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    return ws.max_row  # row 1 is header, so max_row = last data row, next id = max_row

def save_visitor(first_name, last_name, age, visiting):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    next_id = ws.max_row  # header is row 1, so id = current max_row (next data row index)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([next_id, first_name, last_name, int(age), visiting, timestamp])
    # Style data row
    from openpyxl.styles import Font, Alignment, PatternFill
    row = ws.max_row
    fill_color = "F0FFF4" if row % 2 == 0 else "FFFFFF"
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", start_color=fill_color)
    wb.save(EXCEL_FILE)

FORM_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Visitor Registration</title>
    <link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@300;400;600&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

        :root {
            --forest: #1B4332;
            --sage: #2D6A4F;
            --mint: #52B788;
            --pale: #D8F3DC;
            --cream: #F8FDF9;
            --text: #1A2E22;
            --muted: #6B8F71;
            --border: #B7E4C7;
            --white: #FFFFFF;
            --error: #C0392B;
            --success: #1B4332;
        }

        body {
            font-family: 'DM Sans', sans-serif;
            background: var(--cream);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 40px 20px;
        }

        body::before {
            content: '';
            position: fixed;
            inset: 0;
            background:
                radial-gradient(ellipse at 10% 20%, rgba(82,183,136,0.12) 0%, transparent 50%),
                radial-gradient(ellipse at 90% 80%, rgba(27,67,50,0.08) 0%, transparent 50%);
            pointer-events: none;
            z-index: 0;
        }

        .card {
            position: relative;
            z-index: 1;
            background: var(--white);
            border-radius: 20px;
            box-shadow: 0 4px 6px rgba(27,67,50,0.04), 0 20px 60px rgba(27,67,50,0.10);
            width: 100%;
            max-width: 520px;
            overflow: hidden;
        }

        .card-header {
            background: linear-gradient(135deg, var(--forest) 0%, var(--sage) 100%);
            padding: 44px 48px 36px;
            position: relative;
            overflow: hidden;
        }

        .card-header::after {
            content: '';
            position: absolute;
            bottom: -30px;
            right: -30px;
            width: 140px;
            height: 140px;
            border-radius: 50%;
            border: 2px solid rgba(255,255,255,0.08);
        }

        .card-header::before {
            content: '';
            position: absolute;
            top: -20px;
            left: -20px;
            width: 100px;
            height: 100px;
            border-radius: 50%;
            border: 2px solid rgba(255,255,255,0.06);
        }

        .badge {
            display: inline-block;
            background: rgba(255,255,255,0.15);
            color: rgba(255,255,255,0.85);
            font-size: 11px;
            font-weight: 500;
            letter-spacing: 2px;
            text-transform: uppercase;
            padding: 5px 14px;
            border-radius: 20px;
            margin-bottom: 16px;
            border: 1px solid rgba(255,255,255,0.2);
        }

        h1 {
            font-family: 'Cormorant Garamond', serif;
            font-size: 36px;
            font-weight: 300;
            color: var(--white);
            line-height: 1.1;
            letter-spacing: -0.5px;
        }

        h1 span {
            display: block;
            font-weight: 600;
        }

        .card-body {
            padding: 40px 48px 48px;
        }

        .flash-msg {
            padding: 14px 18px;
            border-radius: 10px;
            margin-bottom: 28px;
            font-size: 14px;
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 10px;
        }

        .flash-success {
            background: var(--pale);
            color: var(--success);
            border: 1px solid var(--border);
        }

        .flash-error {
            background: #FDECEA;
            color: var(--error);
            border: 1px solid #F5C6C2;
        }

        .form-row {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 16px;
        }

        .field {
            margin-bottom: 22px;
        }

        label {
            display: block;
            font-size: 11px;
            font-weight: 500;
            letter-spacing: 1.5px;
            text-transform: uppercase;
            color: var(--muted);
            margin-bottom: 8px;
        }

        input, select {
            width: 100%;
            padding: 14px 16px;
            border: 1.5px solid var(--border);
            border-radius: 10px;
            font-family: 'DM Sans', sans-serif;
            font-size: 15px;
            color: var(--text);
            background: var(--cream);
            outline: none;
            transition: border-color 0.2s, box-shadow 0.2s, background 0.2s;
            appearance: none;
        }

        input:focus, select:focus {
            border-color: var(--mint);
            background: var(--white);
            box-shadow: 0 0 0 3px rgba(82,183,136,0.12);
        }

        input::placeholder {
            color: #AECBB4;
        }

        .select-wrap {
            position: relative;
        }

        .select-wrap::after {
            content: '';
            position: absolute;
            right: 16px;
            top: 50%;
            transform: translateY(-50%);
            width: 0;
            height: 0;
            border-left: 5px solid transparent;
            border-right: 5px solid transparent;
            border-top: 6px solid var(--muted);
            pointer-events: none;
        }

        .divider {
            height: 1px;
            background: linear-gradient(to right, transparent, var(--border), transparent);
            margin: 28px 0;
        }

        button[type="submit"] {
            width: 100%;
            padding: 16px;
            background: linear-gradient(135deg, var(--forest) 0%, var(--sage) 100%);
            color: var(--white);
            border: none;
            border-radius: 12px;
            font-family: 'DM Sans', sans-serif;
            font-size: 15px;
            font-weight: 500;
            letter-spacing: 0.3px;
            cursor: pointer;
            transition: opacity 0.2s, transform 0.15s, box-shadow 0.2s;
            box-shadow: 0 4px 20px rgba(27,67,50,0.25);
            position: relative;
            overflow: hidden;
        }

        button[type="submit"]::after {
            content: '';
            position: absolute;
            inset: 0;
            background: rgba(255,255,255,0);
            transition: background 0.2s;
        }

        button[type="submit"]:hover {
            opacity: 0.92;
            transform: translateY(-1px);
            box-shadow: 0 8px 28px rgba(27,67,50,0.30);
        }

        button[type="submit"]:active {
            transform: translateY(0);
        }

        .view-link {
            text-align: center;
            margin-top: 22px;
        }

        .view-link a {
            color: var(--muted);
            font-size: 13px;
            text-decoration: none;
            border-bottom: 1px dashed var(--border);
            transition: color 0.2s;
        }

        .view-link a:hover { color: var(--sage); }

        @media (max-width: 480px) {
            .card-header, .card-body { padding-left: 28px; padding-right: 28px; }
            .form-row { grid-template-columns: 1fr; gap: 0; }
            h1 { font-size: 30px; }
        }
    </style>
</head>
<body>
    <div class="card">
        <div class="card-header">
            <div class="badge">&#x2713; &nbsp;Visitor Check-In</div>
            <h1>Welcome. <span>Sign in below.</span></h1>
        </div>
        <div class="card-body">
            {% with messages = get_flashed_messages(with_categories=true) %}
              {% for category, message in messages %}
                <div class="flash-msg flash-{{ category }}">
                  {{ '✓' if category == 'success' else '!' }} {{ message }}
                </div>
              {% endfor %}
            {% endwith %}

            <form method="POST" action="/submit">
                <div class="form-row">
                    <div class="field">
                        <label for="first_name">First Name</label>
                        <input type="text" id="first_name" name="first_name" placeholder="Jane" required>
                    </div>
                    <div class="field">
                        <label for="last_name">Last Name</label>
                        <input type="text" id="last_name" name="last_name" placeholder="Smith" required>
                    </div>
                </div>

                <div class="field">
                    <label for="age">Age</label>
                    <input type="number" id="age" name="age" placeholder="30" min="1" max="120" required>
                </div>

                <div class="field">
                    <label for="visiting">Who are you visiting?</label>
                    <div class="select-wrap">
                        <select id="visiting" name="visiting" required>
                            <option value="" disabled selected>Select an employee…</option>
                            {% for emp in employees %}
                            <option value="{{ emp }}">{{ emp }}</option>
                            {% endfor %}
                        </select>
                    </div>
                </div>

                <div class="divider"></div>
                <button type="submit">Register Visit &nbsp;→</button>
            </form>

            <div class="view-link">
                <a href="/visitors">View all registered visitors</a>
            </div>
        </div>
    </div>
</body>
</html>
"""

TABLE_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Visitor Log</title>
    <link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@300;600&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: 'DM Sans', sans-serif;
            background: #F8FDF9;
            padding: 40px 24px;
            color: #1A2E22;
        }
        .header {
            max-width: 900px;
            margin: 0 auto 32px;
            display: flex;
            justify-content: space-between;
            align-items: flex-end;
        }
        h1 {
            font-family: 'Cormorant Garamond', serif;
            font-size: 32px;
            font-weight: 300;
            color: #1B4332;
        }
        h1 strong { font-weight: 600; }
        .back-link {
            color: #6B8F71;
            text-decoration: none;
            font-size: 13px;
            border-bottom: 1px dashed #B7E4C7;
        }
        .back-link:hover { color: #2D6A4F; }
        .table-wrap {
            max-width: 900px;
            margin: 0 auto;
            background: #fff;
            border-radius: 16px;
            overflow: hidden;
            box-shadow: 0 4px 30px rgba(27,67,50,0.08);
        }
        table { width: 100%; border-collapse: collapse; }
        thead { background: linear-gradient(135deg, #1B4332, #2D6A4F); }
        thead th {
            padding: 16px 18px;
            color: #fff;
            font-size: 11px;
            font-weight: 500;
            letter-spacing: 1.5px;
            text-transform: uppercase;
            text-align: left;
        }
        tbody tr:nth-child(even) { background: #F0FFF4; }
        tbody tr:hover { background: #D8F3DC; transition: background 0.15s; }
        td {
            padding: 14px 18px;
            font-size: 14px;
            color: #1A2E22;
            border-bottom: 1px solid #E9F7EE;
        }
        .empty {
            text-align: center;
            padding: 60px;
            color: #6B8F71;
            font-size: 15px;
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>Visitor <strong>Log</strong></h1>
        <a class="back-link" href="/">← Back to registration</a>
    </div>
    <div class="table-wrap">
        {% if visitors %}
        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th>First Name</th>
                    <th>Last Name</th>
                    <th>Age</th>
                    <th>Visiting</th>
                    <th>Timestamp</th>
                </tr>
            </thead>
            <tbody>
                {% for v in visitors %}
                <tr>
                    <td>{{ v[0] }}</td>
                    <td>{{ v[1] }}</td>
                    <td>{{ v[2] }}</td>
                    <td>{{ v[3] }}</td>
                    <td>{{ v[4] }}</td>
                    <td>{{ v[5] }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% else %}
        <div class="empty">No visitors registered yet.</div>
        {% endif %}
    </div>
</body>
</html>
"""

@app.route('/')
def index():
    init_excel()
    return render_template_string(FORM_HTML, employees=EMPLOYEES)

@app.route('/submit', methods=['POST'])
def submit():
    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    age = request.form.get('age', '').strip()
    visiting = request.form.get('visiting', '').strip()

    if not all([first_name, last_name, age, visiting]):
        flash("All fields are required.", "error")
        return redirect(url_for('index'))

    if not age.isdigit() or not (1 <= int(age) <= 120):
        flash("Please enter a valid age (1–120).", "error")
        return redirect(url_for('index'))

    save_visitor(first_name, last_name, age, visiting)
    flash(f"Welcome, {first_name}! Your visit has been registered.", "success")
    return redirect(url_for('index'))

@app.route('/visitors')
def visitors():
    if not os.path.exists(EXCEL_FILE):
        return render_template_string(TABLE_HTML, visitors=[])
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(cell is not None for cell in row)]
    return render_template_string(TABLE_HTML, visitors=rows)

if __name__ == '__main__':
    app.run(debug=True)